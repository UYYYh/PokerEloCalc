from openpyxl import Workbook, load_workbook
import os
from DatabaseService import execute_sql
import msal
import requests
import json

output_file = r"D:\UYY\Poker Database\Linstead Poker Data.xlsx"
temp_file = r"D:\UYY\Poker Database\Temp Linstead Poker Data.xlsx"

def export_to_excel():
    # File paths
    output_file = r"D:\UYY\Poker Database\Linstead Poker Data.xlsx"
    temp_file = r"D:\UYY\Poker Database\Temp Linstead Poker Data.xlsx"

    try:
        # Query to fetch player ratings and detailed statistics (for the first sheet)
        query_players = """
        SELECT 
            name, 
            rating,
            (SELECT COUNT(DISTINCT game_id) FROM player_poker_game_info WHERE name = player.name) AS total_games,
            (SELECT COUNT(*) FROM player_poker_game_info WHERE name = player.name AND placement = 1) AS first_place_finishes,
            (SELECT AVG(placement) FROM player_poker_game_info WHERE name = player.name) AS average_placement,
            (SELECT COUNT(*) FROM player_poker_game_info WHERE busted_by = player.name) AS total_players_busted
        FROM player
        ORDER BY rating DESC
        """
        player_results = execute_sql(query_players, {}, fetch=True).fetchall()

        # Prepare player data for export, including calculating average players busted
        player_data = []
        for row in player_results:
            name, rating, total_games, first_place_finishes, average_placement, total_players_busted = row
            average_players_busted = total_players_busted / total_games if total_games > 0 else 0
            player_data.append((name, rating, total_games, first_place_finishes, average_placement, total_players_busted, average_players_busted))

        # Query to fetch detailed poker game info (for the second sheet)
        query_game_info = """
        SELECT 
            name, 
            pki.game_id, 
            placement, 
            busted_by, 
            elo_before, 
            elo_change, 
            game_date
        FROM poker_game_db.player_poker_game_info pki
        JOIN poker_game_db.poker_game pg ON pki.game_id = pg.game_id
        ORDER BY game_date, pki.game_id, placement
        """
        game_results = execute_sql(query_game_info, {}, fetch=True).fetchall()
        game_data = [(row[0], row[1], row[2], row[3], row[4], row[5], row[6].strftime("%Y-%m-%d")) for row in game_results]

        # Load existing workbook or create a new one
        workbook = load_workbook(output_file) if os.path.exists(output_file) else Workbook()

        # Remove "Player Ratings" sheet if it exists
        if "Player Ratings" in workbook.sheetnames:
            workbook.remove(workbook["Player Ratings"])

        # Update or create "Player Ratings and Statistics" sheet
        if "Player Ratings and Statistics" in workbook.sheetnames:
            workbook.remove(workbook["Player Ratings and Statistics"])
        sheet_players = workbook.create_sheet("Player Ratings and Statistics")
        sheet_players.append(["Name", "Rating", "Total Games Played", "First Place Finishes", "Average Placement", "Total Players Busted", "Average Players Busted"])
        for row in player_data:
            sheet_players.append(row)

        # Update or create "Poker Game Info" sheet
        if "Poker Game Info" in workbook.sheetnames:
            workbook.remove(workbook["Poker Game Info"])
        sheet_game_info = workbook.create_sheet("Poker Game Info")
        sheet_game_info.append(["Name", "Game ID", "Placement", "Busted By", "Elo Before", "Elo Change", "Game Date"])

        # Add poker game info data with empty rows between different game_ids
        previous_game_id = None
        for row in game_data:
            current_game_id = row[1]
            if previous_game_id is not None and current_game_id != previous_game_id:
                sheet_game_info.append([])  # Add an empty row to separate different games
            sheet_game_info.append(row)
            previous_game_id = current_game_id

        # Save to temporary file first
        workbook.save(temp_file)

        # Replace the original file with the updated one
        os.replace(temp_file, output_file)
        print(f"Player ratings and poker game info exported and overwritten at {output_file}.")

        # Step 2: Upload the file to OneDrive
        upload_to_onedrive(output_file)  # Use output_file instead of temp_file

    finally:
        # Clean up the temporary file if it still exists
        if os.path.exists(temp_file):
            os.remove(temp_file)




def get_access_token_device_flow(client_id, tenant_id):
    authority = f"https://login.microsoftonline.com/{tenant_id}"
    scope = ["https://graph.microsoft.com/.default"]

    # Create a PublicClientApplication for device code flow
    app = msal.PublicClientApplication(client_id=client_id, authority=authority)

    # Initiate device code flow to acquire token
    flow = app.initiate_device_flow(scopes=scope)
    if "user_code" not in flow:
        raise ValueError("Failed to create device flow. Please check your client_id and tenant_id.")

    print(f"Please go to {flow['verification_uri']} and enter the code {flow['user_code']} to authenticate.")

    # Wait for the user to authenticate
    result = app.acquire_token_by_device_flow(flow)

    if "access_token" in result:
        return result['access_token']
    else:
        raise Exception(f"Could not obtain access token: {result.get('error')}, {result.get('error_description')}")

def upload_to_onedrive(file_path):
    # Authentication parameters

    client_id = os.environ.get("PokerEloCalcFileEditorClientID")
    tenant_id = os.environ.get("PokerEloCalcFileEditorTenantID")
    file_id = os.environ.get("PokerEloCalcFileEditorFileID")    
    user_email = os.environ.get("MyImperialEmail")

    # Get an access token via Device Code Flow
    access_token = get_access_token_device_flow(client_id, tenant_id)

    # URL for overwriting the existing file on OneDrive
    graph_api_endpoint = "https://graph.microsoft.com/v1.0"
    url = f"{graph_api_endpoint}/users/{user_email}/drive/items/{file_id}/content"

    headers = {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    }

    # Upload the Excel file to OneDrive
    with open(file_path, "rb") as file_stream:
        response = requests.put(url, headers=headers, data=file_stream)

    if response.status_code in [200, 201]:
        print("File on OneDrive overwritten successfully.")
    else:
        print(f"Failed to overwrite the file on OneDrive: {response.status_code} - {response.text}")

